import schedule
import pandas as pd
from openpyxl import Workbook, load_workbook
from threading import Thread
import sched
import time
from flask import Flask, render_template

pirma = 'sheets/Pirmdiena.xlsx'
otra = 'sheets/Otrdiena.xlsx'
tresa = 'sheets/Tresdiena.xlsx'
ceturta = 'sheets/Ceturtdiena.xlsx'
piekta = 'sheets/Piektdiena.xlsx'
diena = pirma
workbook = load_workbook(filename=diena)
sheet = workbook.active
stundas = 0
dnr = 0


t = Thread(target=sched.run_schedule)
t.start()

def pirm():
    global diena
    diena = pirma

def otr():
    global diena
    diena = otra

def tres():
    global diena
    diena = tresa

def ceturt():
    global diena
    diena = ceturta

def piekt():
    global diena
    diena = piekta

def nakmad():
    if dnr == 0:
        pirm()
    elif dnr == 1:
        otr()
    elif dnr == 2:
        tres()
    elif dnr == 3:
        ceturt()
    elif dnr == 4:
        piekt()

def piesk():
    global stundas
    stundas = stundas + 1
    print(stundas)
    nakmas()

def nulite():
    global stundas
    stundas = 0
    global dnr
    dnr = dnr + 1
    nakmad()

def nulitedivi():
    global stundas
    stundas = 0
    global dnr
    dnr = 0
    nakmad()

def pirunotr():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=6, amount=8)
    sheet.delete_rows(idx=10, amount=8)
    sheet.delete_rows(idx=14, amount=8)
    workbook.save(filename='need.xlsx')

def otruntre():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=1)
    sheet.delete_rows(idx=6, amount=7)
    sheet.delete_rows(idx=8, amount=1)
    sheet.delete_rows(idx=10, amount=7)
    sheet.delete_rows(idx=12, amount=1)
    sheet.delete_rows(idx=14, amount=7)
    workbook.save(filename='need.xlsx')

def treuncet():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=2)
    sheet.delete_rows(idx=6, amount=6)
    sheet.delete_rows(idx=8, amount=2)
    sheet.delete_rows(idx=10, amount=6)
    sheet.delete_rows(idx=12, amount=2)
    sheet.delete_rows(idx=14, amount=6)
    workbook.save(filename='need.xlsx')

def cetunpie():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=3)
    sheet.delete_rows(idx=6, amount=5)
    sheet.delete_rows(idx=8, amount=3)
    sheet.delete_rows(idx=10, amount=5)
    sheet.delete_rows(idx=12, amount=3)
    sheet.delete_rows(idx=14, amount=5)
    workbook.save(filename='need.xlsx')

def pieunses():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=4)
    sheet.delete_rows(idx=6, amount=4)
    sheet.delete_rows(idx=8, amount=4)
    sheet.delete_rows(idx=10, amount=4)
    sheet.delete_rows(idx=12, amount=4)
    sheet.delete_rows(idx=14, amount=4)
    workbook.save(filename='need.xlsx')

def sesunsep():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=5)
    sheet.delete_rows(idx=6, amount=3)
    sheet.delete_rows(idx=8, amount=5)
    sheet.delete_rows(idx=10, amount=3)
    sheet.delete_rows(idx=12, amount=5)
    sheet.delete_rows(idx=14, amount=3)
    workbook.save(filename='need.xlsx')

def sepunast():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=6)
    sheet.delete_rows(idx=6, amount=2)
    sheet.delete_rows(idx=8, amount=6)
    sheet.delete_rows(idx=10, amount=2)
    sheet.delete_rows(idx=12, amount=6)
    sheet.delete_rows(idx=14, amount=2)
    workbook.save(filename='need.xlsx')

def astundev():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=7)
    sheet.delete_rows(idx=6, amount=1)
    sheet.delete_rows(idx=8, amount=7)
    sheet.delete_rows(idx=10, amount=1)
    sheet.delete_rows(idx=12, amount=7)
    sheet.delete_rows(idx=14, amount=1)
    workbook.save(filename='need.xlsx')

def devundes():
    workbook = load_workbook(filename=diena)
    sheet = workbook.active
    sheet.delete_rows(idx=4, amount=8)
    sheet.delete_rows(idx=8, amount=8)
    sheet.delete_rows(idx=12, amount=8)
    workbook.save(filename='need.xlsx')
 
schedule.every(6).seconds.do(piesk)
schedule.every(1).minutes.do(nulite)
schedule.every(6).minutes.do(nulitedivi)

# schedule.every().monday.at("07:20").do(piesk)
# schedule.every().monday.at("08:50").do(piesk)
# schedule.every().monday.at("09:40").do(piesk)
# schedule.every().monday.at("10:30").do(piesk)
# schedule.every().monday.at("11:20").do(piesk)
# schedule.every().monday.at("12:10").do(piesk)
# schedule.every().monday.at("13:00").do(piesk)
# schedule.every().monday.at("13:50").do(piesk)
# schedule.every().monday.at("14:35").do(piesk)
# schedule.every().monday.at("16:20").do(nulite)

# schedule.every().tuesday.at("07:20").do(piesk)
# schedule.every().tuesday.at("08:50").do(piesk)
# schedule.every().tuesday.at("09:40").do(piesk)
# schedule.every().tuesday.at("10:30").do(piesk)
# schedule.every().tuesday.at("11:20").do(piesk)
# schedule.every().tuesday.at("12:10").do(piesk)
# schedule.every().tuesday.at("13:00").do(piesk)
# schedule.every().tuesday.at("13:50").do(piesk)
# schedule.every().tuesday.at("14:35").do(piesk)
# schedule.every().tuesday.at("16:20").do(nulite)

# schedule.every().wednesday.at("07:20").do(piesk)
# schedule.every().wednesday.at("08:50").do(piesk)
# schedule.every().wednesday.at("09:40").do(piesk)
# schedule.every().wednesday.at("10:30").do(piesk)
# schedule.every().wednesday.at("11:20").do(piesk)
# schedule.every().wednesday.at("12:10").do(piesk)
# schedule.every().wednesday.at("13:00").do(piesk)
# schedule.every().wednesday.at("13:50").do(piesk)
# schedule.every().wednesday.at("14:35").do(piesk)
# schedule.every().wednesday.at("16:20").do(nulite)

# schedule.every().thursday.at("07:20").do(piesk)
# schedule.every().thursday.at("08:50").do(piesk)
# schedule.every().thursday.at("09:40").do(piesk)
# schedule.every().thursday.at("10:30").do(piesk)
# schedule.every().thursday.at("11:20").do(piesk)
# schedule.every().thursday.at("12:10").do(piesk)
# schedule.every().thursday.at("13:00").do(piesk)
# schedule.every().thursday.at("13:50").do(piesk)
# schedule.every().thursday.at("14:35").do(piesk)
# schedule.every().thursday.at("16:20").do(nulite)

# schedule.every().friday.at("07:20").do(piesk)
# schedule.every().friday.at("08:50").do(piesk)
# schedule.every().friday.at("09:40").do(piesk)
# schedule.every().friday.at("10:30").do(piesk)
# schedule.every().friday.at("11:20").do(piesk)
# schedule.every().friday.at("12:10").do(piesk)
# schedule.every().friday.at("13:00").do(piesk)
# schedule.every().friday.at("13:50").do(piesk)
# schedule.every().friday.at("14:35").do(piesk)
# schedule.every().friday.at("16:20").do(nulitedivi)

def nakmas():
    if stundas == 1:
        pirunotr()
    elif stundas == 2:
        otruntre()
    elif stundas == 3:
        treuncet()
    elif stundas == 4:
        cetunpie()
    elif stundas == 5:
        pieunses()
    elif stundas == 6:
        sesunsep()
    elif stundas == 7:
        sepunast()
    elif stundas == 8:
        astundev()
    elif stundas >= 9:
        devundes()

app = Flask(__name__)

@app.route('/sheet')
def hello():
    if stundas == 0:
        return render_template('hello.html')
    elif stundas >= 1:
        df = pd.read_excel('need.xlsx')
        return '{} {}' .format(df.to_html(), render_template('hello.html'))

if __name__ == '__main__':
    app.run()
